#! python3
# Update Spreadsheet.py

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

PRICE_UPDATES = {'GARLIC' : 3.07, 'CELERY' : 1.19 , 'Lemon' : 1.27}

for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row = rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
