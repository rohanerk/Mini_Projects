#! python3

#sendDuesReminder.py - Send mails based on payment status in spreadsheet

import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

lastCol = sheet.max_column
latestMonth = sheet.cell(row = 1, column = lastCol).value

# Check each member's payment status

unpaidMembers = {}
for r in range(2,sheet.max_row + 1):
    payment = sheet.cell(row = r ,column = lastCol).value
    if payment != 'paid':
        name = sheet.cell(row = r,column = 1).value
        email = sheet.cell(row = r, column = 2).value
        unpaidMembers[name] = email

# Log in to email account

smtpObj = smtplib.SMTP('smtp.gmail.com',587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_id' , sys.argv[1])

# Send out reminder mails

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid\nDear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" % (latestMonth, name, latestMonth)

    print('Sending mail to %s...' % email)
    sendmailStatus = smtpObj.sendmail('my_email_id',email,body)

    if sendmailStatus != {}:
        print('There was a problem sending mail to %s:%s' % (email,sendmailStatus))

smtpObj.quit()
